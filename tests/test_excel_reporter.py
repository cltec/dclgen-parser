import unittest
from dclgen_parser.parser import Attribute, Table
from dclgen_parser.excel_reporter import ExcelReporter
import os
from openpyxl import load_workbook

class TestExcelReporter(unittest.TestCase):
    def setUp(self):
        self.reporter = ExcelReporter()
        self.test_tables = [
            Table(
                table_name="test_table_1",
                schema_name="test_schema_1",
                attributes=[
                    Attribute(name="id", type="INTEGER", nullable=False),
                    Attribute(name="name", type="VARCHAR", length=255, nullable=True)
                ]
            ),
            Table(
                table_name="test_table_2",
                schema_name="test_schema_2",
                attributes=[
                    Attribute(name="price", type="DECIMAL", precision=10, scale=2, nullable=False)
                ]
            )
        ]
        self.output_file = "test_report.xlsx"

    def tearDown(self):
        if os.path.exists(self.output_file):
            os.remove(self.output_file)

    def test_generate_excel_report(self):
        self.reporter.generate_excel_report(self.test_tables, self.output_file)
        self.assertTrue(os.path.exists(self.output_file))

        # Load the workbook and verify its contents
        workbook = load_workbook(self.output_file)
        sheet = workbook.active

        # Check headers
        expected_headers = ["Table", "Name", "Type", "Length", "Precision", "Scale", "Nullable"]
        actual_headers = [cell.value for cell in sheet[1]]
        self.assertEqual(expected_headers, actual_headers)

        # Check attribute data
        expected_data = [
            ["test_table_1", "id", "INTEGER", "N/A", "N/A", "N/A", "No"],
            ["test_table_1", "name", "VARCHAR", 255, "N/A", "N/A", "Yes"],
            ["test_table_2", "price", "DECIMAL", "N/A", 10, 2, "No"]
        ]

        for row_idx, expected_row in enumerate(expected_data, start=2):
            actual_row = [sheet.cell(row=row_idx, column=col_idx).value for col_idx in range(1, 8)]
            self.assertEqual(expected_row, actual_row)

if __name__ == '__main__':
    unittest.main()
